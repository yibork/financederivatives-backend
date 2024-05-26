# views.py
from django.forms import ValidationError
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import FormView
from wagtail.admin.messages import success, error
from .forms import JobImportForm
from .models import JobPage, JobIndexPage
from openpyxl import load_workbook
from rest_framework.generics import ListAPIView
from .serializers import JobPageSerializer,JobPageAddSerializer
import os
from django.http import JsonResponse
from django.views import View
from django.core.management import call_command
from django.core.files.storage import default_storage
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework import permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth.models import User
from django.utils.text import slugify
from .models import JobPage, JobIndexPage
from django.utils.decorators import method_decorator

# joblistings/views.py

from django.http import JsonResponse
import os
from googleapiclient.discovery import build
from google.oauth2 import service_account
import uuid
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import requests
import json
from wagtail.models import Page
import uuid
from .models import JobPage, JobIndexPage, Company, City, Industry, Category, Currency, ContactPerson, ContactIntern, CoddingLanguages, Languages

#fetch data from indeed
def fetch_indeed_jobs(request):
    file_path = 'indeed_jobs.json'
    url = "https://indeed11.p.rapidapi.com/"
    payload = {
        "search_terms": "Engineer",
        "location": "United States",
        "page": "1"
    }
    headers = {
        "content-type": "application/json",
        "X-RapidAPI-Key": "3cbf12bb91msh67c4b738331dd30p1908c8jsn89e563d7fccb",
        "X-RapidAPI-Host": "indeed11.p.rapidapi.com"
    }
    response = requests.post(url, json=payload, headers=headers)
    data = response.json()
    # Save or overwrite the fetched data to a file
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

    return response.json()
import logging
logger = logging.getLogger(__name__)
from django.core.exceptions import ObjectDoesNotExist

def get_or_create_related_model(model, value, new_value_field):
    """
    Utility function to get or create a related model instance.
    :param model: The model class
    :param value: The existing value or None
    :param new_value_field: The new value field from the request data
    :return: The model instance
    """
    if new_value_field:
        instance, _ = model.objects.get_or_create(name=new_value_field)
    elif value:
        instance = model.objects.get(pk=value)
    else:
        instance = None
    return instance

class JobPageAddViewSet(viewsets.ModelViewSet):
    queryset = JobPage.objects.all()
    serializer_class = JobPageAddSerializer
    permission_classes = [permissions.IsAuthenticated]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        # Handling new entries or lookup existing
        company = None
        if data.get('new_company_name'):
            company, _ = Company.objects.get_or_create(name=data['new_company_name'])
        elif data.get('company_name'):
            try:
                company = data['company_name']
            except ObjectDoesNotExist:
                return Response({'company_name': 'Company does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        location = None
        if data.get('new_city_name'):
            location, _ = City.objects.get_or_create(name=data['new_city_name'])
        elif data.get('location'):
            try:
                location = data['location']
            except ObjectDoesNotExist:
                return Response({'location': 'City does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        industry = None
        if data.get('new_industry_name'):
            industry, _ = Industry.objects.get_or_create(name=data['new_industry_name'])
        elif data.get('industry'):
            try:
                industry = data['industry']
                print(industry)
            except ObjectDoesNotExist:
                return Response({'industry': 'Industry does not exist.'}, status=status.HTTP_400_BAD_REQUEST)
        print('yassine')
        print(industry)

        category = None
        if data.get('new_category_name'):
            category, _ = Category.objects.get_or_create(name=data['new_category_name'])
        elif data.get('categories'):
            try:
                category = data['categories']
            except ObjectDoesNotExist:
                return Response({'categories': 'Category does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        currency = None
        if data.get('new_currency_code'):
            currency, _ = Currency.objects.get_or_create(code=data['new_currency_code'])
        elif data.get('currency'):
            try:
                currency = data['currency']
            except ObjectDoesNotExist:
                return Response({'currency': 'Currency does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        # Use the IDs for both coding_languages and language_requirements
        coding_languages = [CoddingLanguages.objects.get(pk=id) for id in data.get('coding_languages', [])]
        language_requirements = [Languages.objects.get(pk=id) for id in data.get('language_requirements', [])]

        job_index, created = JobIndexPage.objects.get_or_create(
            title="Job Index Page",
            defaults={
                'title': 'Job Index Page',
                'slug': 'job-index-page',
                'depth': 4,  # Set the depth for the index page
                'path': '0001000100010005',  # Set a unique path for the index page
            }
        )

        if created:
            root = Page.get_first_root_node()
            job_index = root.add_child(instance=job_index)
        else:
            job_index = JobIndexPage.objects.get(title="Job Index Page")

        slug = slugify(data['job_title']) + '-' + str(uuid.uuid4())[:8]

        base_path = job_index.path
        depth = job_index.depth + 1
        path_suffix = 1

        while True:
            job_page_path = f"{base_path}{path_suffix:04d}"
            if not JobPage.objects.filter(path=job_page_path).exists():
                break
            path_suffix += 1
        job_page = JobPage(
            title=data['job_title'],
            slug=slug,
            job_title=data['job_title'],
            company_name=company,
            location=location,
            contract_type=data['contract_type'],
            duration=data['duration'],
            month=data['month'],
            job_description=data['job_description'],
            notes_feedbacks=data['notes_feedbacks'],
            expected_salary=data['expected_salary'],
            currency=currency,
            asset_class=data['asset_class'],
            interview_report_link=data['interview_report_link'],
            industry=industry,
            categories=category,
            owner=request.user,
            live=False,
            path=job_page_path,
            depth=depth,
        )

        job_index.add_child(instance=job_page)

        job_page.save()
        job_page.coding_languages.set(coding_languages)
        job_page.language_requirements.set(language_requirements)

        for contact_person_data in data.get('contact_persons', []):
            contact_person = ContactPerson(
                job_page=job_page,
                name=contact_person_data.get('name'),
                linkedin=contact_person_data.get('linkedin'),
                email=contact_person_data.get('email')
            )
            contact_person.save()

        for contact_intern_data in data.get('contact_interns', []):
            contact_intern = ContactIntern(
                job_page=job_page,
                name=contact_intern_data.get('name'),
                linkedin=contact_intern_data.get('linkedin'),
                email=contact_intern_data.get('email')
            )
            contact_intern.save()

        job_page.unpublish()

        return Response({"message": "Job added successfully."}, status=status.HTTP_201_CREATED)

def transform_to_google_format(indeed_jobs):
    google_jobs = []
    uniqId = uuid.UUID
    for job in indeed_jobs:
        google_job = {
            "job": {
                "title": job['job_title'],
                "companyName": job['company_name'],
                "description": job['summary'],
                "addresses": [job['location']],
                "requisitionId": uniqId,  # This should be unique
                "applicationInfo": {
                    "uris": [job['url']]
                },
                "postingPublishTime": "2024-05-10T12:00:00Z",  # Example timestamp
                "postingExpireTime": "2024-06-10T12:00:00Z"
            }
        }
        google_jobs.append(google_job)

    return google_jobs
        
        
def post_indeed_jobs_to_google(request):
    # Fetch jobs from Indeed (assuming `fetch_indeed_jobs()` is already implemented)
    file_path = 'indeed_jobs.json'  # The path to the JSON file where the jobs are stored
    
    try:
        # Try to open and read the JSON file
        with open(file_path, 'r') as file:
            indeed_jobs = json.load(file)
        return JsonResponse({'Jobs': indeed_jobs})
    except FileNotFoundError:
        # If the file is not found, perhaps handle it or return an error
        return JsonResponse({'error': 'Job data file not found'}, status=404)
    except json.JSONDecodeError:
        # Handle cases where the JSON data is corrupted or improperly formatted
        return JsonResponse({'error': 'Error decoding JSON data'}, status=500)
    # rest of the code for case of storing data in
    google_jobs = transform_to_google_format(indeed_jobs)
    body = {"jobs": google_jobs}
    return JsonResponse({'Jobs': google_jobs})
    # Post jobs to Google Cloud Talent Solution
    #response = client_service.projects().tenants().jobs().batchCreate(
    #    parent='projects/dashboard-finance-derivatives/tenants/fahd-g7o51', 
    #    body=body
    #).execute()
    #print('Response: ', response)


# Configure your Google Cloud credentials
'''
dir_path = os.path.dirname(os.path.realpath(__file__))
SERVICE_ACCOUNT_FILE =  os.path.join(dir_path, 'secret', 'dashboard-finance-derivatives-0b16f2f20dc3.json')
SCOPES = ['https://www.googleapis.com/auth/jobs']
print("Attempting to load credentials from:", SERVICE_ACCOUNT_FILE)
credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)

client_service = build('jobs', 'v3', credentials=credentials)
project_id = 'projects/dashboard-finance-derivatives'
        
@csrf_exempt
@require_http_methods(["POST"])  # Ensuring only POST requests are handled
def job_search(request):
    if request.method != 'POST':
        return JsonResponse({'message': 'Invalid request method, POST required'}, status=405)
    # Hardcoding the job type in the query
    search_term = 'Engineer'  # Default to 'Engineer' if not specified
    job_categories = ['SCIENCE_AND_ENGINEERING']  # Example category, adjust as necessary

    request_metadata = {
    'domain': 'yourcompanydomain.com',
    'session_id': str(request.session.session_key),  # Assuming session is being used
    'user_id': str(request.user.id) if request.user.is_authenticated else 'anonymous'
    }

    job_query = {
        'query': search_term,
        'jobCategories': job_categories  # Use jobCategories instead of jobType
    }

    body = {
        'searchMode': 'JOB_SEARCH',
        'requestMetadata': request_metadata,
        'jobQuery': job_query,
        'jobView': 'JOB_VIEW_FULL',
        'pageSize': 20
    }
    print('Request body: ', body)
    response = client_service.projects().jobs().search(parent=project_id, body=body).execute()
    if 'error' in response:
        print('Error response: ', response['error'])
        return JsonResponse({'message': 'Error processing your request', 'details': response['error']}, status=500)

    print('API response: ', response)
    jobs = response.get('matchingJobs', [])
    print(jobs)
    if not jobs:
        return JsonResponse({'message': 'No Job Results'}, status=404)

    results = [{
        'title': job.get('job', {}).get('title'),
        'description': job.get('job', {}).get('description')
    } for job in jobs]

    return JsonResponse({'jobs': results})
'''
class ExcelUploadView(View):
    def get(self, request, *args, **kwargs):
        # Render the upload form template
        return render(request, 'admin/excel_upload_form.html')

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return JsonResponse({'error': 'No file uploaded.'}, status=400)

        file_path = default_storage.save('tmp/excel_file.xlsx', excel_file)
        full_file_path = default_storage.path(file_path)

        try:
            # Call your management command here
            call_command('import_jobs', full_file_path)
            return JsonResponse({'success': 'Jobs imported successfully.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        finally:
            # Clean up the file after import
            os.remove(full_file_path)

class JobImportView(FormView):
    template_name = 'job_import.html'
    form_class = JobImportForm
    success_url = reverse_lazy('wagtailadmin_home')  # Redirect to Wagtail admin home

    def form_valid(self, form):
        try:
            excel_file = form.cleaned_data['excel_file']
            wb = load_workbook(filename=excel_file)
            ws = wb.active
            job_index_page = JobIndexPage.objects.live().first()

            if not job_index_page:
                return error(self.request, "Job Index Page does not exist.")

            for row in ws.iter_rows(min_row=2, values_only=True):
                job_page_data = {
                    'title': row[0],
                    'job_title': row[0],
                    'company_name': row[1],
                    'location': row[2],
                    # ... other fields ...
                }
                job_page = JobPage(**job_page_data)
                job_index_page.add_child(instance=job_page)
                job_page.save_revision().publish()

            success(self.request, "Jobs imported successfully.")
        except Exception as e:
            error(self.request, f"An error occurred: {e}")
            return super().form_invalid(form)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # ... additional context data setup ...
        return context
class JobImportView(FormView):
    template_name = 'job_import.html'
    form_class = JobImportForm
    success_url = reverse_lazy('wagtailadmin_home')  # Redirect to Wagtail admin home

    def form_valid(self, form):
        try:
            excel_file = form.cleaned_data['excel_file']
            wb = load_workbook(filename=excel_file)
            ws = wb.active
            job_index_page = JobIndexPage.objects.live().first()

            if not job_index_page:
                return error(self.request, "Job Index Page does not exist.")

            for row in ws.iter_rows(min_row=2, values_only=True):
                job_page_data = {
                    'title': row[0],
                    'job_title': row[0],
                    'company_name': row[1],
                    'location': row[2],
                    # ... other fields ...
                }
                job_page = JobPage(**job_page_data)
                job_index_page.add_child(instance=job_page)
                job_page.save_revision().publish()

            success(self.request, "Jobs imported successfully.")
        except Exception as e:
            error(self.request, f"An error occurred: {e}")
            return super().form_invalid(form)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # ... additional context data setup ...
        return context
class JobPageViewSet(viewsets.ModelViewSet):
    serializer_class = JobPageSerializer

    def get_queryset(self):
        """
        Optionally filter the pages to those that are public based on a condition.
        For example, the condition can be based on a query parameter or a setting.
        """
        queryset = JobPage.objects.live()

        # Here we check a condition - for example, checking a query parameter
        if 'public' in self.request.query_params:
            # Only return public pages if 'public' query parameter is provided
            return queryset.public().order_by('-first_published_at')
        else:
            return queryset.all()