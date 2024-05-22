from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from django.core.exceptions import ObjectDoesNotExist
from wagtail.models import Site
from joblistings.models import JobPage, JobIndexPage, Company, City
import datetime

class Command(BaseCommand):
    help = 'Import jobs from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='The path to the Excel file')

    def handle(self, *args, **options):
        file_path = options['file_path']
        wb = load_workbook(filename=file_path)
        ws = wb.active

        # Ensure that there is a JobIndexPage
        try:
            job_index_page = JobIndexPage.objects.live().first()
            if not job_index_page:
                raise ObjectDoesNotExist("Job Index Page does not exist.")
        except ObjectDoesNotExist:
            job_index_page = self.create_job_index_page()
            self.stdout.write(self.style.SUCCESS('Job Index Page created successfully.'))

        # Import jobs
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            self.import_job_row(row_index, row, job_index_page)

    def import_job_row(self, row_index, row, job_index_page):
        # Unpack all the row values
        (job_title, company_name, location, industry, categories, contract_type, 
         duration, start_date, job_description, notes_feedbacks, language_requirements, 
         expected_salary, coding_languages, asset_class, interview_report_link, *rest) = row

        # Ensure that the mandatory fields are not empty (for example, job_title)
        if not job_title:
            self.stdout.write(self.style.WARNING(f'Skipping job with empty title at row {row_index}'))
            return

        # Get or create the related objects
        company = self.get_or_create_company(company_name)
        city = self.get_or_create_city(location)

        # Prepare the job page data dictionary
        job_page_data = {
            'title': job_title.strip() if job_title else "No Title",
            'job_title': job_title,
            'company_name': company,
            'location': city,
            'industry': industry,
            'categories': categories,
            'contract_type': contract_type,
            'duration': duration,
            'start_date': self.parse_date(start_date),
            'job_description': job_description,
            'notes_feedbacks': notes_feedbacks,
            'language_requirements': language_requirements,
            'expected_salary': expected_salary,
            'coding_languages': coding_languages,
            'asset_class': asset_class,
            'interview_report_link': interview_report_link,
        }

        try:
            job_page = JobPage(**job_page_data)
            job_index_page.add_child(instance=job_page)
            job_page.save_revision().publish()
            self.stdout.write(self.style.SUCCESS(f'Job "{job_title}" imported successfully at row {row_index}.'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error saving job "{job_title}" at row {row_index}: {e}'))

    def create_job_index_page(self):
        root = Site.objects.get(is_default_site=True).root_page
        job_index_page = JobIndexPage(title="Main Job Listings")
        root.add_child(instance=job_index_page)
        job_index_page.save_revision().publish()
        return job_index_page

    def get_or_create_company(self, company_name):
        if company_name:
            company, created = Company.objects.get_or_create(name=company_name)
            if created:
                self.stdout.write(self.style.SUCCESS(f'Created new company: {company_name}'))
            return company
        return None

    def get_or_create_city(self, city_name):
        if city_name:
            city, created = City.objects.get_or_create(name=city_name)
            if created:
                self.stdout.write(self.style.SUCCESS(f'Created new city: {city_name}'))
            return city
        return None

    def parse_date(self, date_string):
        """
        Parse the date from the format in the Excel sheet to a Python date object.
        You may need to adjust the parsing depending on the date format.
        """
        if date_string:
            try:
                return datetime.datetime.strptime(date_string, '%m/%d/%Y').date()
            except ValueError:
                self.stdout.write(self.style.WARNING(f'Invalid date format: {date_string}'))
        return None
